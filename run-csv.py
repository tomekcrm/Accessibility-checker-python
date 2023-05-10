# Importing required modules
import os
import csv
import openpyxl
from selenium import webdriver
from axe_selenium_python import Axe
from selenium.webdriver.firefox.options import Options
from openpyxl.utils import get_column_letter
import zipfile
from datetime import datetime
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders

# Function for retrieving reports from Axe tool
def get_violations(page):
    options = Options()
    options.add_argument('-headless')
    driver = webdriver.Firefox(options=options)
    driver.get(page)
    axe = Axe(driver)
    axe.inject()
    results = axe.run()
    driver.quit()
    return results["violations"]

# Function for saving report to a CSV file
def save_report(page, violations):
    with open(f"pages/{page.split('/')[-1]}.csv", 'w', newline='') as f:
        writer = csv.writer(f)
        writer.writerow(["Page", "Impact", "Description"])
        for violation in violations:
            writer.writerow([page, violation['impact'], violation['description']])

def send_email(file_path, email_to):
    email_from = os.getenv('EMAIL_ADDRESS')
    password = os.getenv('EMAIL_PASSWORD')
    smtp_server = os.getenv('SMTP_SERVER')
    smtp_port = os.getenv('SMTP_PORT')
    
    msg = MIMEMultipart()
    msg['From'] = email_from
    msg['To'] = email_to
    msg['Subject'] = f"Accessibility Report {datetime.now().strftime('%Y-%m-%d')}"

    with open(file_path, 'rb') as f:
        part = MIMEBase('application', 'zip')
        part.set_payload(f.read())
        encoders.encode_base64(part)
        part.add_header('Content-Disposition', f'attachment; filename="{os.path.basename(file_path)}"')
        msg.attach(part)

    # Add list of checked pages to email body
    pages_checked = []
    wb = openpyxl.load_workbook('results.xlsx')
    ws = wb.active
    ws.cell(row=1, column=2, value='Page')
    for i in range(2, ws.max_row + 1):
        page = ws.cell(row=i, column=1).value
        pages_checked.append(page)

    body = "The following pages were checked:\n\n" + "\n".join(pages_checked)
    msg.attach(MIMEText(body, "plain"))

    with smtplib.SMTP(smtp_server, smtp_port) as server:
        server.starttls()
        server.login(email_from, password)
        server.sendmail(email_from, email_to, msg.as_string())

# Main function
def main():
    # Opening a CSV file with pages data
    with open('pages.csv', newline='') as csvfile:
        reader = csv.reader(csvfile)
        header = next(reader)
        pages = [row[0] for row in reader]
        
    total_pages = len(pages)
    
    # Initializing worksheet with headers and data
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.cell(row=1, column=1, value='Page')
    ws.cell(row=1, column=2, value='Link')
    
    # Getting violations and saving results to worksheet
    error_types = set()
    for i, page in enumerate(pages, start=2):
        ws.cell(row=i, column=1, value=page)
        ws.cell(row=i, column=2, value=f'=HYPERLINK("pages/{page.split("/")[-1]}.csv", "Link to report")')
        violations = get_violations(page)
        save_report(page, violations)
        for violation in violations:
            error_types.add(violation['impact'])
        for j, error_type in enumerate(sorted(error_types), start=3):
            ws.cell(row=1, column=j, value=error_type)
        for j, error_type in enumerate(sorted(error_types), start=3):
            count = sum(1 for v in violations if v['impact'] == error_type)
            ws.cell(row=i, column=j, value=count)
        print(f'Testing page {i-1}/{total_pages}: {page}')
    ws.column_dimensions[get_column_letter(2)].width = 20
    wb.save('results.xlsx')

    # Zipping the results.xlsx file and the pages directory into an archive
    pages_folder = 'pages'
    zip_filename = f'report_{total_pages}_pages_created_{datetime.now().strftime("%Y-%m-%d_%H-%M-%S")}.zip'
    with zipfile.ZipFile(zip_filename, 'w', zipfile.ZIP_DEFLATED) as zipf:
        zipf.write('results.xlsx')
        for folder, subfolders, files in os.walk(pages_folder):
            for file in files:
                zipf.write(os.path.join(folder, file))

    # Sending the ZIP file to the specified email address (optional)
    send_report = input("Do you want to send the report to the e-mail address provided? (Y/N): ")
    if send_report.lower() == 'y':
        email_to = input("Enter the e-mail address to which you want to send the report: ")
        send_email(zip_filename, email_to)
        print("Report send, you can find it locally too")
    else:
        print("The report has been saved locally.")

if __name__ == '__main__':
    main()
