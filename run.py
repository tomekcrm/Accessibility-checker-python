# Importing required modules
import os
import openpyxl
from selenium import webdriver
from axe_selenium_python import Axe
from selenium.webdriver.firefox.options import Options
from openpyxl.utils import get_column_letter
import zipfile
from datetime import datetime
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders

# Function for retrieving reports from Axe tool
def get_violations(page):
    options = Options()
    options.add_argument('-headless')
    driver = webdriver.Firefox(options=options)
    driver.get(page)
    axe = Axe(driver)
    axe.inject()
    results = axe.run()
    driver.quit()
    return results["violations"]

# Function for saving report to a text file
def save_report(page, violations):
    with open(f"pages/{page.split('/')[-1]}.txt", 'w') as f:
        f.write(f"Report for page: {page}\n")
        for violation in violations:
            f.write(f"{violation['impact']}: {violation['description']}\n")

def send_email(file_path, email_to):
    email_from = os.getenv('EMAIL_ADDRESS')
    password = os.getenv('EMAIL_PASSWORD')
    smtp_server = os.getenv('SMTP_SERVER')
    smtp_port = os.getenv('SMTP_PORT')
    
    msg = MIMEMultipart()
    msg['From'] = email_from
    msg['To'] = email_to
    msg['Subject'] = f"Accessibility Report {datetime.now().strftime('%Y-%m-%d')}"

    with open(file_path, 'rb') as f:
        part = MIMEBase('application', 'zip')
        part.set_payload(f.read())
        encoders.encode_base64(part)
        part.add_header('Content-Disposition', f'attachment; filename="{os.path.basename(file_path)}"')
        msg.attach(part)

    # Add list of checked pages to email body
    pages_checked = []
    wb = openpyxl.load_workbook('results.xlsx')
    ws = wb.active
    ws.cell(row=1, column=2, value='Page')
    for i in range(2, ws.max_row + 1):
        page = ws.cell(row=i, column=1).value
        pages_checked.append(page)

    body = "The following pages were checked:\n\n" + "\n".join(pages_checked)
    msg.attach(MIMEText(body, "plain"))

    with smtplib.SMTP(smtp_server, smtp_port) as server:
        server.starttls()
        server.login(email_from, password)
        server.sendmail(email_from, email_to, msg.as_string())

# Main function
def main():
    # Opening a workbook with pages data
    wb = openpyxl.load_workbook('pages.xlsx')
    ws = wb.active
    ws.cell(row=1, column=2, value='Page')
    error_types = set()
    for i in range(2, ws.max_row + 1):
        # Retrieving the page URL from the worksheet and adding it to the 'Page' column
        page = ws.cell(row=i, column=1).value
        ws.cell(row=i, column=2, value=page)
        # Retrieving the report from the Axe tool
        violations = get_violations(page)
        save_report(page, violations)
        # Counting the number of errors and adding headers to the worksheet
        for violation in violations:
            error_types.add(violation['impact'])
        for j, error_type in enumerate(sorted(error_types), start=4):
            ws.cell(row=1, column=j, value=error_type)
        # Counting the number of errors and saving them in the remaining columns
        for j, error_type in enumerate(sorted(error_types), start=4):
            count = sum(1 for v in violations if v['impact'] == error_type)
            ws.cell(row=i, column=j, value=count)
        # Adding a link to the report for each page in the 3rd column
        page_name = ws.cell(row=i, column=2).value.split('/')[-1]
        link = f'=HYPERLINK("pages/{page_name}.txt", "Link to report")'
        ws.cell(row=i, column=3, value=link)
        print(f'Testing page {i-1}/{ws.max_row-1}: {page}')
    ws.cell(row=1, column=3, value='Link')
    ws.column_dimensions[get_column_letter(3)].width = 20
    wb.save('results.xlsx')


    # Zipping the results.xlsx file and the pages directory into an archive
    pages_folder = 'pages'
    zip_filename = f'report_{ws.max_row - 1}_pages___created_{datetime.now().strftime("%Y-%m-%d_%H-%M-%S")}.zip'
    with zipfile.ZipFile(zip_filename, 'w', zipfile.ZIP_DEFLATED) as zipf:
        zipf.write('results.xlsx')
        for folder, subfolders, files in os.walk(pages_folder):
            for file in files:
                zipf.write(os.path.join(folder, file))

    # Sending the ZIP file to the specified email address (optional)
    send_report = input("Do you want to send the report to the e-mail address provided? (Y/N): ")
    if send_report.lower() == 'y':
        email_to = input("Enter the e-mail address to which you want to send the report: ")
        send_email(zip_filename, email_to)
        print("report sended, you can find it locally too")
    else:
        print("The report has been saved locally.")


if __name__ == '__main__':
    main()
